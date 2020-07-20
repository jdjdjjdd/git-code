from common.venv.var import *
from openpyxl import Workbook,load_workbook


def edit_exc(t1,t2,t3,t4,t5,t7,t6=None,t8=None):
    wb = None
    try:
        if not os.path.exists(myLocalFile_advance):
            wb = Workbook()
            ws = wb.create_sheet('Sheet1')
        else:
            wb = load_workbook(myLocalFile_advance)
            ws = wb['Sheet1']
            wb.remove(ws)
            ws = wb.create_sheet('Sheet1')
        value1 = ['姓名','身份证号码','工号','入职日期','在职状态','离职/转正/自离日期','上班天数','备注']
        value2 = [t1,t2,t3,t4,t5,t6,t7,t8]
        for i in range(1,9):
            ws.cell(1, i).value = value1[i-1]
        for i in range(1,9):
            ws.cell(2, i).value = value2[i-1]
    finally:
        if wb:
            wb.save(myLocalFile_advance)


def edit_exctwo(names,idcards,workcards,realpay,work_entry_date,workhours,workstatus='在职',workstatedate=None,remark=None):
    wb = None
    try:
        if not os.path.exists(myLocalFile_month):
            wb = Workbook()
            ws = wb.create_sheet('Sheet1')
        else:
            wb = load_workbook(myLocalFile_month)
            ws = wb['Sheet1']
            wb.remove(ws)
            ws = wb.create_sheet('Sheet1')
        value1 = ['姓名','身份证号码','工号','实发工资','入职日期','在职状态','离职/转正日期','备注','出勤小时数']
        for i in range(1,10):
            ws.cell(1, i).value = value1[i-1]

        n = len(names)
        for j in range(2,n+2):
            ws.cell(j,1).value = names[j-2]
            ws.cell(j,2).value = idcards[j-2]
            ws.cell(j,3).value = workcards[j-2]
            ws.cell(j,4).value = realpay
            ws.cell(j,5).value = work_entry_date
            ws.cell(j,6).value = workstatus
            ws.cell(j,7).value = workstatedate
            ws.cell(j,8).value = remark
            ws.cell(j,9).value = workhours

    finally:
        if wb:
            wb.save(myLocalFile_month)


